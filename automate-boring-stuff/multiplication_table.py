#! python3
# multiplication_table.py - Takes a number N from the command line and
# creates an N×N multiplication table in an Excel spreadsheet

import sys, openpyxl, os
from openpyxl.styles import Font

desktop = os.path.join(os.path.expanduser('~'), 'Desktop')
os.chdir(desktop)

def multiplication_table(number):
    wb = openpyxl.Workbook()
    sheet = wb.active

    # Label Line 1 
    for col in range(1, number + 1):
        cell = sheet.cell(row=1, column=col + 1, value=col)
        cell.font = Font(bold=True)

    # Label column A
    for row in range(1, number + 1):
        cell = sheet.cell(row=row + 1, column=1, value=row)
        cell.font = Font(bold=True)

    for row in range(1, number + 1):
        for col in range(1, number + 1):
            sheet.cell(row=row + 1, column=col + 1, value=row * col)

    wb.save(f'multiplication_table_{number}.xlsx')
    print('Done')

num = 0
if len(sys.argv) > 1:
    num = int(sys.argv[1])
else:
    print('you should write a number to perfom the table')
    sys.exit()

multiplication_table(num)
