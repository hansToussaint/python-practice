#! python3
# read_census_excel.py - Tabulates population and number of census tracts for each county.

import openpyxl, pprint, os

print('Opening workbook...')

desktop = os.path.join(os.path.expanduser('~'), 'Desktop')
wb = openpyxl.load_workbook(os.path.join(desktop, 'censuspopdata.xlsx'))
sheet = wb['Population by Census Tract']

county_data = {}

# Fill in country_data with each county's population and tracts.
print('Reading rows...')
for row in range(2, sheet.max_row + 1):
    # Each row in the spreadsheet has data for one census tract.
    #state = sheet['B' + str(row)].value
    state = sheet.cell(row=row, column=2).value
    #county = sheet['C' + str(row)].value
    county = sheet.cell(row=row, column=3).value
    #pop = sheet['D' + str(row)].value
    pop = sheet.cell(row=row, column=4).value

    # Make sure the key for this state exists.
    county_data.setdefault(state, {})

    # Make sure the key for this county in this state exists.
    county_data[state].setdefault(county, {'tracts': 0, 'pop': 0})

    #each row represents one census tract, so increment by one.
    county_data[state][county]['tracts'] += 1
    # Increase the county pop by the pop in this census tract.
    county_data[state][county]['pop'] += int(pop)

# Open a new text file and write the contents of county_data
print('Writing results...')
with open(os.path.join(desktop, 'census2010.py'), 'w') as f:
    f.write('all_data= ' + pprint.pformat(county_data))

print('Done.')
