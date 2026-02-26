#! python3
#update_produce_excel.py - Corrects costs in produce sales spreadsheet

import openpyxl, os

desktop = os.path.join(os.path.expanduser('~'), 'Desktop')
wb = openpyxl.load_workbook(os.path.join(desktop, 'produceSales.xlsx'))
sheet = wb['Sheet']

# The produce tyoes and their uodated prices
PRICE_UPDATES = {'Garlic':3.07,
                 'Celery': 1.19,
                 'Lemon': 1.27
                 }

#Loop through the rows and update the prices
for row_num in range(2, sheet.max_row + 1): # skip the first row
    produce_name = sheet.cell(row=row_num, column=1).value
    
    if produce_name in PRICE_UPDATES:
        sheet.cell(row=row_num, column=2).value = PRICE_UPDATES[produce_name]

wb.save(os.path.join(desktop, 'updatedProduceSales.xlsx'))
